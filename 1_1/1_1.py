# In the name of Allah
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook


def derivative(x, y):
    dy = y[1:] - y[:-1]
    dx = x[1:] - x[:-1]

    ydot = y.copy()
    ydot[1:] = dy / dx
    ydot[0] = ydot[1]
    return ydot


ws = load_workbook('input_function.xlsx', data_only=True)['Sheet1']
data = []
for i in range(1,  ws.max_row):
    data.append([ws[f'A{i}'].value, ws[f'B{i}'].value])
data = np.array(data)

x = data[:, 0]
y = data[:, 1]

ydot = derivative(x, y)
y2dot = derivative(x, ydot)

wb_save = Workbook()
ws_save = wb_save.active

for i in range(len(x)):
    ws_save[f'A{i + 1}'] = x[i]
    ws_save[f'B{i + 1}'] = ydot[i]
    ws_save[f'C{i + 1}'] = y2dot[i]

wb_save.save('derivatives.xlsx')

plt.plot(x, y, label='exp')
plt.plot(x, ydot, label='1st derivative')
plt.plot(x, y2dot, label='2nd derivative')
plt.xlabel('x')
plt.ylabel('y')
plt.title('Exp with its 1st & 2nd derivatives')
plt.legend()
plt.show()
